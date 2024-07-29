import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Function to save message to Excel file
def save_to_excel(name, email, message):
    excel_file = 'messages.xlsx'

    # Check if Excel file exists
    if not os.path.exists(excel_file):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Messages'
        header = ['ID', 'Name', 'Email', 'Message', 'Timestamp']
        sheet.append(header)
    else:
        workbook = load_workbook(excel_file)

        # Check if 'Messages' worksheet exists
        if 'Messages' not in workbook.sheetnames:
            sheet = workbook.create_sheet(title='Messages')
            header = ['ID', 'Name', 'Email', 'Message', 'Timestamp']
            sheet.append(header)
        else:
            sheet = workbook['Messages']

    # Append new message data
    row = sheet.max_row + 1 if sheet.max_row else 1
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data = [row, name, email, message, timestamp]
    sheet.append(data)

    # Save workbook back to Excel file
    workbook.save(excel_file)
